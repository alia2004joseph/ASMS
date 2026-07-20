"""
Excel exporter for the Reports module.
"""

from __future__ import annotations

import io
from collections.abc import Mapping, Sequence
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from reports.exporters.base_exporter import BaseExporter


class ExcelExporter(BaseExporter):
    """
    Export report data as an Excel workbook (.xlsx).
    """

    format_code = "EXCEL"
    content_type = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    file_extension = ".xlsx"

    def render(
        self,
        *,
        context: Mapping[str, Any],
    ) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active

        worksheet.title = str(
            context.get("sheet_name", "Report")
        )[:31]

        headers = context.get("headers", ())
        self._write_headers(worksheet, headers)

        groups = context.get("groups")
        if groups:
            self._write_groups(worksheet, groups, len(headers))
        else:
            self._write_rows(worksheet, context.get("rows", ()))

        self._write_grand_total(
            worksheet,
            context.get("grand_total"),
        )

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def _write_headers(worksheet, headers: Sequence[Any]) -> None:
        if not headers:
            return
        worksheet.append(list(headers))
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

    @staticmethod
    def _write_rows(worksheet, rows: Sequence[Sequence[Any]]) -> None:
        for row in rows:
            worksheet.append(list(row))

    @staticmethod
    def _write_groups(
        worksheet,
        groups: Sequence[Mapping[str, Any]],
        header_count: int,
    ) -> None:
        for group in groups:
            worksheet.append(
                [group.get("label", "")] + [""] * max(header_count - 1, 0)
            )

            for row in group.get("rows", ()):
                worksheet.append(list(row))

            subtotal = group.get("subtotal")
            if subtotal is not None:
                worksheet.append(["Subtotal", *subtotal])
                for cell in worksheet[worksheet.max_row]:
                    cell.font = Font(italic=True)

    @staticmethod
    def _write_grand_total(
        worksheet,
        grand_total: Sequence[Any] | None,
    ) -> None:
        if grand_total is None:
            return

        worksheet.append(["Grand Total", *grand_total])

        for cell in worksheet[worksheet.max_row]:
            cell.font = Font(bold=True)